# accounts/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.forms import ModelForm
from accounts.models import Perfil
import re
import logging

from api.forms import CargaMasivaForm
from api.models import CalificacionTributaria

logger = logging.getLogger(__name__)


# ===============================
# Vista de Registro CORREGIDA
# ===============================
def registro(request):
    """Vista de registro con validación completa"""
    
    if request.method == "POST":
        # Obtener datos del formulario
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')
        
        logger.info(f"📝 Intento de registro - Username: {username}, Email: {email}")
        
        # Lista de errores
        errors = []
        
        # ============ VALIDACIONES ============
        
        # 1. Validar username
        if not username:
            errors.append("El nombre de usuario es obligatorio")
        elif len(username) < 3:
            errors.append("El nombre de usuario debe tener al menos 3 caracteres")
        elif not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', username):
            errors.append("El nombre de usuario debe comenzar con letra y solo contener letras, números y guiones bajos")
        elif User.objects.filter(username=username).exists():
            errors.append("Este nombre de usuario ya está en uso")
        
        # 2. Validar email
        if not email:
            errors.append("El correo electrónico es obligatorio")
        elif not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            errors.append("Ingresa un correo electrónico válido")
        elif User.objects.filter(email=email).exists():
            errors.append("Este correo electrónico ya está registrado")
        
        # 3. Validar contraseña
        if not password:
            errors.append("La contraseña es obligatoria")
        elif len(password) < 8:
            errors.append("La contraseña debe tener al menos 8 caracteres")
        elif not re.search(r'[A-Z]', password):
            errors.append("La contraseña debe contener al menos una letra mayúscula")
        elif not re.search(r'[a-z]', password):
            errors.append("La contraseña debe contener al menos una letra minúscula")
        elif not re.search(r'\d', password):
            errors.append("La contraseña debe contener al menos un número")
        elif not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
            errors.append("La contraseña debe contener al menos un carácter especial")
        
        # 4. Validar confirmación
        if password != password2:
            errors.append("Las contraseñas no coinciden")
        
        # ============ SI HAY ERRORES ============
        if errors:
            logger.warning(f"❌ Registro fallido - Errores: {errors}")
            for error in errors:
                messages.error(request, error)
            return render(request, "accounts/registro.html")
        
        # ============ CREAR USUARIO ============
        try:
            # Crear el usuario
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            
            logger.info(f"✅ Usuario creado en BD: {user.username} (ID: {user.id})")
            
            # Verificar que se creó el perfil automáticamente
            try:
                perfil = user.perfil
                logger.info(f"✅ Perfil creado automáticamente: ID={perfil.id}, aprobado={perfil.aprobado}")
            except Perfil.DoesNotExist:
                # Si no existe, crearlo manualmente
                logger.warning(f"⚠️ Perfil no creado automáticamente - Creando manualmente...")
                perfil = Perfil.objects.create(user=user, aprobado=False)
                logger.info(f"✅ Perfil creado manualmente: ID={perfil.id}")
            
            messages.success(
                request,
                "¡Cuenta creada correctamente! Tu cuenta será revisada por un administrador antes de ser activada."
            )
            logger.info(f"✅ Registro completado exitosamente para: {username}")
            
            return redirect("accounts:login")
            
        except Exception as e:
            logger.error(f"❌ ERROR al crear usuario: {str(e)}", exc_info=True)
            messages.error(request, f"Error al crear la cuenta. Por favor, intenta de nuevo.")
            return render(request, "accounts/registro.html")
    
    else:
        # GET - Mostrar formulario vacío
        logger.debug("📄 Mostrando formulario de registro (GET)")
        return render(request, "accounts/registro.html")


# ===============================
# Vistas de Autenticación Web
# ===============================
def mi_login(request):
    """Vista de login con validación de aprobación"""
    
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        
        if form.is_valid():
            user = form.get_user()
            
            # ============ VALIDACIÓN DE APROBACIÓN ============
            try:
                perfil = user.perfil
                
                # Verificar si el perfil está aprobado
                if not perfil.aprobado:
                    logger.warning(
                        f"⚠️ Intento de login de usuario no aprobado - "
                        f"Usuario: {user.username} | IP: {request.META.get('REMOTE_ADDR')}"
                    )
                    messages.error(
                        request,
                        "Tu cuenta está pendiente de aprobación por un administrador. "
                        "Por favor, espera a que tu cuenta sea activada."
                    )
                    return redirect("accounts:login")
                
            except Perfil.DoesNotExist:
                logger.error(
                    f"❌ Usuario sin perfil intentó iniciar sesión - Usuario: {user.username}"
                )
                messages.error(
                    request,
                    "Tu cuenta tiene un problema. Contacta al administrador."
                )
                return redirect("accounts:login")
            
            # Si llegamos aquí, el usuario está aprobado
            login(request, user)
            logger.info(
                f"✅ Login exitoso - Usuario: {user.username} | IP: {request.META.get('REMOTE_ADDR')}"
            )
            messages.success(request, f"¡Bienvenido, {user.username}!")
            return redirect("accounts:home")
        else:
            username = request.POST.get('username', 'desconocido')
            logger.warning(
                f"❌ Login fallido (credenciales incorrectas) - "
                f"Usuario: {username} | IP: {request.META.get('REMOTE_ADDR')}"
            )
            messages.error(request, "Usuario o contraseña incorrectos.")
    else:
        form = AuthenticationForm()
    
    return render(request, "login.html", {"form": form})

@login_required
def mi_logout(request):
    username = request.user.username
    logout(request)
    logger.info(f"🚪 Logout - Usuario: {username}")
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect("accounts:login")


# ===============================
# Dashboard / Home
# ===============================
@login_required
def home(request):
    """Dashboard con estadísticas y widget de conversión"""
    from api.models import CargaMasiva
    
    logger.debug(f"🏠 Acceso a home - Usuario: {request.user.username}")
    
    # Obtener estadísticas
    total_calificaciones = CalificacionTributaria.objects.filter(usuario=request.user).count()
    total_cargas = CargaMasiva.objects.filter(iniciado_por=request.user).count()
    ultimas_cargas = CargaMasiva.objects.filter(iniciado_por=request.user).order_by('-fecha_inicio')[:3]
    
    context = {
        'total_calificaciones': total_calificaciones,
        'total_cargas': total_cargas,
        'ultimas_cargas': ultimas_cargas,
    }
    
    logger.debug(f"   Calificaciones: {total_calificaciones}, Cargas: {total_cargas}")
    
    return render(request, "home.html", context)


# ===============================
# Cargas Masivas
# ===============================
@login_required
def lista_cargas(request):
    """Lista todas las cargas masivas del usuario"""
    from api.models import CargaMasiva
    
    logger.info(f"📋 Lista de cargas - Usuario: {request.user.username}")
    
    try:
        cargas = CargaMasiva.objects.filter(iniciado_por=request.user).order_by('-fecha_inicio')
        logger.debug(f"   Total de cargas: {cargas.count()}")
        
        return render(request, "accounts/lista_cargas.html", {"cargas": cargas})
        
    except Exception as e:
        logger.error(f"❌ Error listando cargas: {str(e)}", exc_info=True)
        messages.error(request, f"Error al cargar la lista: {str(e)}")
        return render(request, "accounts/lista_cargas.html", {"cargas": []})


@login_required
def nueva_carga(request):
    """Vista para crear nueva carga masiva con información de columnas"""
    from api.excel_handler import ExcelHandler, get_columnas_esperadas
    
    logger.info(f"{'='*60}")
    logger.info(f"🔵 INICIO nueva_carga() - Usuario: {request.user.username}")
    logger.info(f"   Método: {request.method}")

    if request.method == "POST":
        logger.info(f"📤 POST detectado")
        logger.info(f"   FILES recibidos: {list(request.FILES.keys())}")
        logger.info(f"   POST data keys: {list(request.POST.keys())}")
        
        form = CargaMasivaForm(request.POST, request.FILES)
        
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            tipo_carga = form.cleaned_data["tipo_carga"]
            mercado = form.cleaned_data["mercado"]
            
            logger.info(f"✅ Formulario válido")
            logger.info(f"   - Archivo: {archivo.name}")
            logger.info(f"   - Tamaño: {archivo.size} bytes ({archivo.size/1024:.2f} KB)")
            logger.info(f"   - Content-Type: {archivo.content_type}")
            logger.info(f"   - Tipo de carga: {tipo_carga}")
            logger.info(f"   - Mercado: {mercado}")

            try:
                logger.info(f"🔄 Iniciando procesamiento con ExcelHandler...")
                handler = ExcelHandler(archivo, tipo_carga, request.user)
                handler.mercado = mercado
                
                logger.info(f"🔄 Ejecutando handler.procesar()...")
                carga = handler.procesar()
                
                logger.info(f"✅ ¡CARGA EXITOSA! ID: {carga.id}")
                logger.info(f"   - Estado: {carga.estado}")
                logger.info(f"   - Registros procesados: {carga.registros_procesados}")
                logger.info(f"   - Registros exitosos: {carga.registros_exitosos}")
                logger.info(f"   - Registros fallidos: {carga.registros_fallidos}")
                
                messages.success(request, f"Carga procesada correctamente (ID {carga.id}). Exitosos: {carga.registros_exitosos}, Fallidos: {carga.registros_fallidos}")
                return redirect("accounts:detalle_carga", pk=carga.id)

            except Exception as e:
                logger.error(f"❌ ERROR CRÍTICO en procesamiento")
                logger.error(f"   Tipo de error: {type(e).__name__}")
                logger.error(f"   Mensaje: {str(e)}")
                logger.error(f"   Traceback completo:", exc_info=True)
                
                messages.error(request, f"Error procesando carga: {str(e)}")
        else:
            logger.error(f"❌ Formulario INVÁLIDO")
            logger.error(f"   Errores: {form.errors}")
            logger.error(f"   Errores no de campo: {form.non_field_errors()}")
            
            messages.error(request, f"Formulario inválido: {form.errors}")

    else:
        form = CargaMasivaForm()
        logger.info(f"📋 Mostrando formulario vacío")

    # Obtener columnas esperadas según tipo de carga
    columnas_basicas = get_columnas_esperadas('MONITOR')
    columnas_factores = get_columnas_esperadas('FACTORES')
    
    context = {
        'form': form,
        'columnas_basicas': columnas_basicas,
        'columnas_factores': columnas_factores,
    }

    logger.info(f"{'='*60}")
    return render(request, "accounts/nueva_carga.html", context)


@login_required
def detalle_carga(request, pk):
    from api.models import CargaMasiva
    
    logger.info(f"👁️ Accediendo a detalle_carga - ID: {pk} - Usuario: {request.user.username}")
    
    try:
        carga = get_object_or_404(CargaMasiva, pk=pk, iniciado_por=request.user)
        calificaciones = CalificacionTributaria.objects.filter(carga_masiva=carga)
        
        logger.info(f"✅ Carga encontrada: {carga}")
        logger.info(f"   Calificaciones asociadas: {calificaciones.count()}")
        
        return render(request, "accounts/detalle_carga.html", {
            "carga": carga, 
            "calificaciones": calificaciones
        })
        
    except Exception as e:
        logger.error(f"❌ Error obteniendo detalle de carga {pk}: {str(e)}", exc_info=True)
        raise


@login_required
def descargar_plantilla(request):
    """Genera y descarga una plantilla Excel con dos hojas: Datos y Factores"""
    logger.info(f"📥 Descarga de plantilla - Usuario: {request.user.username}")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # ============================================
        # HOJA 1: DATOS PRINCIPALES
        # ============================================
        ws1 = wb.active
        ws1.title = "Datos Principales"
        
        # Headers principales
        headers_principales = [
            'corredor_dueno',
            'rut_es_el_manual',
            'ano_comercial',
            'mercado',
            'instrumento',
            'fecha_pago',
            'secuencia_evento',
            'numero_dividendo',
            'descripcion',
            'tipo_sociedad',
            'divisa',
            'acopio_lsfxf',
            'valor_historico',
            'origen',
            'es_local'
        ]
        
        # Estilo para headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Escribir headers
        for col, header in enumerate(headers_principales, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos de ejemplo
        ejemplo_principales = [
            'Corredor ABC',           # corredor_dueno
            '12345678-9',             # rut_es_el_manual
            '2025',                   # ano_comercial
            'LOCAL',                  # mercado
            'BONOS-2025',             # instrumento
            '2025-12-31',             # fecha_pago
            '1',                      # secuencia_evento
            '1',                      # numero_dividendo
            'Ejemplo de calificación', # descripcion
            'SA',                     # tipo_sociedad
            'CLP',                    # divisa
            'FALSE',                  # acopio_lsfxf
            '1500000.00',             # valor_historico
            'Manual',                 # origen
            'TRUE'                    # es_local
        ]
        
        for col, valor in enumerate(ejemplo_principales, 1):
            ws1.cell(row=2, column=col, value=valor)
        
        # Ajustar ancho de columnas
        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 3, 20)  # Máximo 20
            ws1.column_dimensions[column].width = adjusted_width
        
        # ============================================
        # HOJA 2: FACTORES
        # ============================================
        ws2 = wb.create_sheet(title="Factores")
        
        # Headers de factores
        ws2.cell(row=1, column=1, value="ID_Registro")
        ws2.cell(row=1, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=1).alignment = header_alignment
        
        # Headers de factores 8-37
        for i in range(8, 38):
            col_num = i - 7 + 1  # Columna 2 para factor_8, 3 para factor_9, etc.
            cell = ws2.cell(row=1, column=col_num)
            cell.value = f'factor_{i}'
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ejemplo de datos de factores
        ws2.cell(row=2, column=1, value="1")  # ID_Registro (debe coincidir con la fila de Datos Principales)
        
        # Factores de ejemplo (puedes poner valores o dejarlos vacíos)
        factores_ejemplo = [
            '1.05',   # factor_8
            '1.03',   # factor_9
            '0.98',   # factor_10
            '1.00',   # factor_11
            '1.02',   # factor_12
            '0.99',   # factor_13
            '1.01',   # factor_14
            '1.04',   # factor_15
            '',       # factor_16
            '',       # factor_17
            '',       # factor_18
            '',       # factor_19
            '',       # factor_20
            '',       # factor_21
            '',       # factor_22
            '',       # factor_23
            '',       # factor_24
            '',       # factor_25
            '',       # factor_26
            '',       # factor_27
            '',       # factor_28
            '',       # factor_29
            '',       # factor_30
            '',       # factor_31
            '',       # factor_32
            '',       # factor_33
            '',       # factor_34
            '',       # factor_35
            '',       # factor_36
            ''        # factor_37
        ]
        
        for col_num, valor in enumerate(factores_ejemplo, 2):  # Empieza en columna 2
            ws2.cell(row=2, column=col_num, value=valor)
        
        # Ajustar ancho de columnas de factores
        ws2.column_dimensions['A'].width = 15
        for i in range(2, 33):  # Columnas B a AF (factores)
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
        
        # ============================================
        # HOJA 3: INSTRUCCIONES
        # ============================================
        ws3 = wb.create_sheet(title="Instrucciones")
        
        # Título
        ws3.cell(row=1, column=1, value="INSTRUCCIONES DE USO")
        ws3.cell(row=1, column=1).font = Font(bold=True, size=14, color="4472C4")
        
        instrucciones = [
            "",
            "HOJA 'Datos Principales':",
            "- Contiene la información básica de cada calificación tributaria",
            "- Llene una fila por cada calificación",
            "- Los campos obligatorios son: corredor_dueno, instrumento, mercado, divisa",
            "",
            "HOJA 'Factores':",
            "- Contiene los factores de actualización (factor_8 hasta factor_37)",
            "- ID_Registro debe coincidir con el número de fila en 'Datos Principales'",
            "- Los factores son opcionales, deje en blanco si no aplica",
            "",
            "FORMATOS:",
            "- Fechas: YYYY-MM-DD (ejemplo: 2025-12-31)",
            "- Números decimales: Use punto como separador (ejemplo: 1500000.00)",
            "- Booleanos: TRUE o FALSE",
            "",
            "DIVISAS SOPORTADAS:",
            "USD, CLP, EUR, COP, PEN, MXN, BRL, ARS",
            "",
            "MERCADOS:",
            "- LOCAL",
            "- INTERNACIONAL",
            "",
            "NOTAS:",
            "- No modifique los nombres de las columnas",
            "- No elimine las hojas de la plantilla",
            "- Guarde el archivo en formato .xlsx"
        ]
        
        for row_num, texto in enumerate(instrucciones, 2):
            cell = ws3.cell(row=row_num, column=1)
            cell.value = texto
            if texto.startswith("HOJA") or texto.startswith("FORMATOS") or texto.startswith("DIVISAS") or texto.startswith("MERCADOS") or texto.startswith("NOTAS"):
                cell.font = Font(bold=True, size=11, color="4472C4")
        
        ws3.column_dimensions['A'].width = 80
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=plantilla_carga_nuam_completa.xlsx"
        
        logger.info(f"✅ Plantilla completa generada correctamente (3 hojas)")
        return response
        
    except Exception as e:
        logger.error(f"❌ Error generando plantilla: {str(e)}", exc_info=True)
        messages.error(request, f"Error al generar plantilla: {str(e)}")
        return redirect("accounts:nueva_carga")

# ===============================
# Calificaciones
# ===============================
@login_required
def lista_calificaciones(request):
    logger.debug(f"📋 Listando calificaciones - Usuario: {request.user.username}")
    calificaciones = CalificacionTributaria.objects.filter(usuario=request.user)
    logger.debug(f"   Total encontradas: {calificaciones.count()}")
    return render(request, "accounts/lista_calificaciones.html", {"calificaciones": calificaciones})


@login_required
def detalle_calificacion(request, pk):
    logger.debug(f"👁️ Detalle calificación {pk} - Usuario: {request.user.username}")
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, usuario=request.user)
    return render(request, "accounts/detalle_calificacion.html", {"calificacion": calificacion})


@login_required
def editar_calificacion(request, pk):
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, usuario=request.user)

    class CalificacionForm(ModelForm):
        class Meta:
            model = CalificacionTributaria
            fields = "__all__"

    if request.method == "POST":
        form = CalificacionForm(request.POST, instance=calificacion)
        if form.is_valid():
            form.save()
            logger.info(f"✏️ Calificación {pk} actualizada por {request.user.username}")
            messages.success(request, "Calificación actualizada correctamente.")
            return redirect("accounts:lista_calificaciones")
        else:
            logger.warning(f"❌ Error actualizando calificación {pk}: {form.errors}")
    else:
        form = CalificacionForm(instance=calificacion)

    return render(request, "accounts/editar_calificacion.html", {"form": form, "calificacion": calificacion})


@login_required
def eliminar_calificacion(request, pk):
    calificacion = get_object_or_404(CalificacionTributaria, pk=pk, usuario=request.user)

    if request.method == "POST":
        calificacion.delete()
        logger.info(f"🗑️ Calificación {pk} eliminada por {request.user.username}")
        messages.success(request, "Calificación eliminada correctamente.")
        return redirect("accounts:lista_calificaciones")

    return render(request, "accounts/eliminar_calificacion.html", {"calificacion": calificacion})


# ===============================
# Logs del Sistema
# ===============================
@login_required
def lista_logs(request):
    """Muestra los logs de operaciones del usuario"""
    from api.models import LogOperacion
    
    logger.debug(f"📋 Lista de logs - Usuario: {request.user.username}")
    
    try:
        logs = LogOperacion.objects.filter(usuario=request.user).order_by('-fecha_hora')[:100]
        logger.debug(f"   Total de logs: {logs.count()}")
        
        return render(request, "accounts/lista_logs.html", {"logs": logs})
        
    except Exception as e:
        logger.error(f"❌ Error listando logs: {str(e)}", exc_info=True)
        return render(request, "accounts/lista_logs.html", {"logs": []})